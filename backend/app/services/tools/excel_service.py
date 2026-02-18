"""Excel to CSV conversion service."""

import io
from typing import Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

from app.core.exceptions import BadRequestError, FileProcessingError


class ExcelService:
    """Service for Excel file processing."""

    async def to_csv(
        self,
        excel_bytes: bytes,
        sheets: Optional[list[str]] = None,
        preserve_formulas: bool = False,
        clean_data: bool = True,
    ) -> list[Tuple[str, bytes, int, int]]:
        """
        Convert Excel to CSV.
        Returns list of (sheet_name, csv_bytes, row_count, col_count).
        """
        try:
            # Load workbook for sheet info
            wb = load_workbook(io.BytesIO(excel_bytes), data_only=not preserve_formulas)

            available_sheets = wb.sheetnames

            if sheets:
                # Validate requested sheets exist
                for sheet in sheets:
                    if sheet not in available_sheets:
                        raise BadRequestError(
                            message=f"Sheet '{sheet}' not found. Available sheets: {', '.join(available_sheets)}"
                        )
                sheets_to_process = sheets
            else:
                sheets_to_process = available_sheets

            results = []

            for sheet_name in sheets_to_process:
                # Read sheet with pandas for better handling
                df = pd.read_excel(
                    io.BytesIO(excel_bytes),
                    sheet_name=sheet_name,
                    engine="openpyxl",
                )

                if clean_data:
                    # Remove completely empty rows and columns
                    df = df.dropna(how="all")
                    df = df.dropna(axis=1, how="all")

                    # Strip whitespace from string columns
                    for col in df.select_dtypes(include=["object"]).columns:
                        df[col] = df[col].apply(
                            lambda x: x.strip() if isinstance(x, str) else x
                        )

                # Get row/col counts
                row_count, col_count = df.shape

                # Convert to CSV
                csv_buffer = io.StringIO()
                df.to_csv(csv_buffer, index=False)
                csv_bytes = csv_buffer.getvalue().encode("utf-8")

                results.append((sheet_name, csv_bytes, row_count, col_count))

                # If preserve_formulas, also create a formulas column version
                if preserve_formulas:
                    formula_df = self._get_formulas(excel_bytes, sheet_name)
                    if formula_df is not None:
                        formula_csv = io.StringIO()
                        formula_df.to_csv(formula_csv, index=False)
                        formula_bytes = formula_csv.getvalue().encode("utf-8")
                        results.append((
                            f"{sheet_name}_formulas",
                            formula_bytes,
                            formula_df.shape[0],
                            formula_df.shape[1],
                        ))

            return results

        except BadRequestError:
            raise
        except Exception as e:
            raise FileProcessingError(message=f"Excel conversion failed: {str(e)}")

    async def get_info(self, excel_bytes: bytes) -> dict:
        """Get information about an Excel file."""
        try:
            wb = load_workbook(io.BytesIO(excel_bytes), read_only=True)

            sheets_info = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Get dimensions
                min_row = ws.min_row or 1
                max_row = ws.max_row or 1
                min_col = ws.min_column or 1
                max_col = ws.max_column or 1

                row_count = max_row - min_row + 1
                col_count = max_col - min_col + 1

                sheets_info.append({
                    "name": sheet_name,
                    "rows": row_count,
                    "columns": col_count,
                })

            wb.close()

            return {
                "sheet_count": len(sheets_info),
                "sheets": sheets_info,
            }

        except Exception as e:
            raise FileProcessingError(message=f"Failed to read Excel file: {str(e)}")

    async def csv_to_excel(
        self,
        csv_bytes: bytes,
        sheet_name: str = "Sheet1",
        delimiter: str = ",",
    ) -> Tuple[bytes, int, int]:
        """
        Convert CSV to Excel.
        Returns (excel_bytes, row_count, col_count).
        """
        try:
            # Detect encoding
            import chardet
            detected = chardet.detect(csv_bytes)
            encoding = detected.get("encoding", "utf-8") or "utf-8"

            # Read CSV
            try:
                df = pd.read_csv(
                    io.BytesIO(csv_bytes),
                    delimiter=delimiter,
                    encoding=encoding,
                )
            except UnicodeDecodeError:
                # Fallback to latin-1
                df = pd.read_csv(
                    io.BytesIO(csv_bytes),
                    delimiter=delimiter,
                    encoding="latin-1",
                )

            row_count, col_count = df.shape

            # Write to Excel
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            excel_bytes = excel_buffer.getvalue()

            return excel_bytes, row_count, col_count

        except Exception as e:
            raise FileProcessingError(message=f"CSV to Excel conversion failed: {str(e)}")

    def _get_formulas(self, excel_bytes: bytes, sheet_name: str) -> Optional[pd.DataFrame]:
        """Extract formulas from a sheet."""
        try:
            # Load without data_only to get formulas
            wb = load_workbook(io.BytesIO(excel_bytes), data_only=False)
            ws = wb[sheet_name]

            # Check if sheet has any formulas
            has_formulas = False
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        has_formulas = True
                        break
                if has_formulas:
                    break

            if not has_formulas:
                return None

            # Extract data with formulas
            data = []
            for row in ws.iter_rows():
                row_data = []
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        row_data.append(f"FORMULA: {cell.value}")
                    else:
                        row_data.append(cell.value)
                data.append(row_data)

            if not data:
                return None

            # Create DataFrame
            df = pd.DataFrame(data[1:], columns=data[0] if data else None)
            return df

        except Exception:
            return None
